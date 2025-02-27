import os
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

from fpdf import FPDF
from datetime import datetime
from collections import Counter

def generate_expense_report(valid_receipts, invalid_receipts, output_path, user_inputs):
    """
    Generates an expense report in Excel and PDF formats with:
    - A summary sheet with compliance statistics.
    - Detailed sheets for compliant and non-compliant expenses.
    - A breakdown of all receipts and their individual items.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    excel_path = generate_excel_report(valid_receipts, invalid_receipts, f"{output_path}.xlsx", user_inputs)
    pdf_path = generate_pdf_report(valid_receipts, invalid_receipts, f"{output_path}.pdf", user_inputs)
    
    return excel_path, pdf_path

def generate_excel_report(valid_receipts, invalid_receipts, output_path, user_inputs):
    """
    Generates an expense report based on the provided template.
    """
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_PATH = os.path.join(script_dir, "assets", "Template.xlsx")
    
    # Load the Excel template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    sheet = wb.active

    # Set header values
    sheet["B3"] = datetime.today().strftime('%Y-%m-%d')  # Submission Date
    sheet["B7"] = f"{user_inputs.get('travel_start_date', 'Not Provided')} to {user_inputs.get('travel_end_date', 'Not Provided')}"  # Travel Dates
    sheet["D3"], sheet["E3"] = user_inputs.get("requester", ""), user_inputs.get("requester_department", "")
    sheet["D5"], sheet["E5"] = user_inputs.get("approver", ""), user_inputs.get("approver_department", "")
    sheet["D7"], sheet["E7"] = user_inputs.get("client", ""), user_inputs.get("project", "")

    # Define starting row for expense entries
    START_ROW = 11
    current_row = START_ROW

    for receipt in valid_receipts + invalid_receipts:
        sheet.cell(row=current_row, column=1, value=receipt["category"])  # Category
        sheet.cell(row=current_row, column=2, value=receipt["receipt_id"])  # Reference
        sheet.cell(row=current_row, column=3, value=receipt["merchant"])  # Details
        sheet.cell(row=current_row, column=4, value=receipt["receipt_id"])  # Invoice
        sheet.cell(row=current_row, column=5, value=receipt["total"])  # Amount $
        sheet.cell(row=current_row, column=6, value="✅ Compliant" if receipt.get("is_compliant", False) else "❌ Non-Compliant")  # Compliance Status
        sheet.cell(row=current_row, column=7, value="; ".join(receipt.get("violations", [])))  # Compliance Violations
        
        current_row += 1  # Move to the next row

    output_file_path = os.path.join(output_path)
    wb.save(output_file_path)
    print(f"✅ Excel report generated: {output_file_path}")

    return output_file_path

def generate_pdf_report(valid_receipts, invalid_receipts, pdf_path, user_inputs):
    """
    Generates a visually enhanced PDF expense report.
    """
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(script_dir, "assets", "logo.png")
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Add Cover Page
    pdf.set_font("Arial", "B", 24)
    pdf.cell(200, 40, txt="Expense Report", ln=True, align="C")
    pdf.set_font("Arial", "I", 16)
    pdf.cell(200, 10, txt=f"Generated on: {current_date}", ln=True, align="C")
    pdf.image(logo_path, x=55, y=100, w=100)
    pdf.add_page()

    # Add Details
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, txt="Report Details", ln=True)
    pdf.set_font("Arial", size=12)
    
    pdf.cell(200, 8, txt=f"Travel Dates: {user_inputs.get('travel_start_date', 'N/A')} to {user_inputs.get('travel_end_date', 'N/A')}", ln=True)
    pdf.cell(200, 8, txt=f"Requester: {user_inputs.get('requester', 'N/A')} ({user_inputs.get('requester_department', 'N/A')})", ln=True)
    pdf.cell(200, 8, txt=f"Approver: {user_inputs.get('approver', 'N/A')} ({user_inputs.get('approver_department', 'N/A')})", ln=True)
    pdf.cell(200, 8, txt=f"Client: {user_inputs.get('client', 'N/A')} | Project: {user_inputs.get('project', 'N/A')}", ln=True)
    pdf.ln(10)

    # Add Summary Section
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, txt="Summary", ln=True)
    pdf.set_font("Arial", size=12)

    total_receipts = len(valid_receipts) + len(invalid_receipts)
    total_valid = len(valid_receipts)
    total_invalid = len(invalid_receipts)

    pdf.cell(200, 10, txt=f"Total Receipts Processed: {total_receipts}", ln=True)
    pdf.cell(200, 10, txt=f"Total Compliant: {total_valid}", ln=True)
    pdf.cell(200, 10, txt=f"Total Non-Compliant: {total_invalid}", ln=True)
    pdf.ln(10)

    # Add Pie Chart for Compliant vs Non-Compliant
    labels = ["Compliant", "Non-Compliant"]
    sizes = [total_valid, total_invalid]
    colors = ["#85BF49", "#FE4E48"]

    plt.figure()
    plt.pie(sizes, labels=labels, colors=colors, autopct="%1.1f%%", startangle=90)
    plt.axis("equal")
    plt.title("Compliant vs Non-Compliant Expenses")
    chart_path = "compliance_pie_chart.png"
    plt.savefig(chart_path)
    plt.close()

    pdf.image(chart_path, x=50, y=pdf.get_y(), w=100)
    pdf.ln(100)

    # Add Top Violations
    most_common_violations = Counter(
        [violation for r in invalid_receipts for violation in r["violations"]]
    ).most_common(5)

    pdf.set_font("Arial", "B", 14)
    pdf.cell(200, 10, txt="Top Violations", ln=True)
    pdf.set_font("Arial", size=12)

    for idx, (violation, count) in enumerate(most_common_violations):
        pdf.cell(200, 10, txt=f"{idx+1}. {violation} ({count} times)", ln=True)
    pdf.ln(10)

    # Add Compliant Expenses
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, txt="Compliant Expenses", ln=True)
    pdf.set_font("Arial", size=12)

    for r in valid_receipts:
        pdf.cell(200, 10, txt=f"Merchant: {r.get('merchant', 'Unknown')}", ln=True)
        pdf.cell(200, 10, txt=f"Date: {r.get('date', 'Unknown')}", ln=True)
        pdf.cell(200, 10, txt=f"Total Amount: ${r.get('total', 0.0):.2f}", ln=True)
        pdf.cell(200, 10, txt=f"Category: {r.get('category', 'Other')}", ln=True)
        pdf.ln(5)
        
        # Add Receipt Breakdown Table
        pdf.set_font("Arial", "B", 12)
        pdf.cell(100, 8, txt="Item Name", border=1)
        pdf.cell(50, 8, txt="Price ($)", border=1)
        pdf.ln()
        pdf.set_font("Arial", size=12)

        for item in r.get("items", []):
            pdf.cell(100, 8, txt=item["name"], border=1)
            pdf.cell(50, 8, txt=f"${item['price']:.2f}", border=1)
            pdf.ln()
        
        pdf.ln(10)

    # Add Non-Compliant Expenses
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 10, txt="Non-Compliant Expenses", ln=True)
    pdf.set_font("Arial", size=12)

    for r in invalid_receipts:
        pdf.cell(200, 10, txt=f"Merchant: {r.get('merchant', 'Unknown')}", ln=True)
        pdf.cell(200, 10, txt=f"Date: {r.get('date', 'Unknown')}", ln=True)
        pdf.cell(200, 10, txt=f"Total Amount: ${r.get('total', 0.0):.2f}", ln=True)
        pdf.cell(200, 10, txt=f"Category: {r.get('category', 'Other')}", ln=True)
        
        violations = r.get("violations", [])
        if violations:
            pdf.cell(200, 10, txt="Violations:", ln=True)
            pdf.set_font("Arial", size=12)
            for v in violations:
                pdf.cell(200, 8, txt=f"- {v}", ln=True)
        
        pdf.ln(5)
        
        # Add Receipt Breakdown Table
        pdf.set_font("Arial", "B", 12)
        pdf.cell(100, 8, txt="Item Name", border=1)
        pdf.cell(50, 8, txt="Price ($)", border=1)
        pdf.ln()
        pdf.set_font("Arial", size=12)

        for item in r.get("items", []):
            pdf.cell(100, 8, txt=item["name"], border=1)
            pdf.cell(50, 8, txt=f"${item['price']:.2f}", border=1)
            pdf.ln()

        pdf.ln(10)

    # Save PDF
    pdf.output(pdf_path)